import openpyxl


class HomePageData:

    test_HomePage_data = [{"name": "Truong", "password": "111111", "email": "truong@gmail.com", "gender": "Male", "bday": "01/01/1999"},
                          {"name": "Thuong", "password": "222222", "email": "thuong@hotmail.com", "gender": "Female", "bday": "11/11/2000"}]

    @staticmethod
    def get_test_data(test_case_name):
        dict_test_data = {}
        book = openpyxl.load_workbook("D:\Course\PythonDemo.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1, column=2)
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "TestCase02":
                for j in range(2, sheet.max_column + 1):
                    dict_test_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        
        print "Git test 1"

        return [dict_test_data]


